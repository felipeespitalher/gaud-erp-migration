"""Tests for Phase 2 parsers: CSV, Excel, Access."""
import pytest
import tempfile
from pathlib import Path

from src.parser.csv_parser import CsvParser
from src.parser.excel_parser import ExcelParser
from src.parser.access_parser import AccessParser
from src.parser.parser_factory import BackupParserFactory


class TestCsvParser:
    """Test CSV parser."""

    def test_csv_parse_simple(self):
        """Test parsing simple CSV."""
        csv_content = """name,age,city
John,30,New York
Jane,25,Los Angeles
Bob,35,Chicago"""

        parser = CsvParser()
        schema = parser.parse(csv_content)

        assert len(schema.tables) == 1
        table = schema.tables[0]
        assert table.name == "data"
        assert len(table.columns) == 3
        assert table.columns[0].name == "name"
        assert table.columns[1].name == "age"
        assert table.columns[2].name == "city"
        assert table.estimated_rows == 3

    def test_csv_parse_with_semicolon(self):
        """Test parsing CSV with semicolon delimiter."""
        csv_content = """product;price;quantity
Notebook;2500.00;10
Mouse;50.00;100
Keyboard;150.00;50"""

        parser = CsvParser()
        schema = parser.parse(csv_content)

        assert len(schema.tables) == 1
        table = schema.tables[0]
        assert len(table.columns) == 3
        assert table.estimated_rows == 3

    def test_csv_type_inference(self):
        """Test type inference in CSV."""
        csv_content = """id,name,price,active
1,Product A,99.99,true
2,Product B,149.99,false
3,Product C,199.99,true"""

        parser = CsvParser()
        schema = parser.parse(csv_content)

        table = schema.tables[0]
        # Check if types are inferred reasonably
        # id should be INTEGER
        assert table.columns[0].type == "INTEGER"
        # name should be VARCHAR
        assert table.columns[1].type == "VARCHAR"
        # price should be FLOAT
        assert table.columns[2].type == "FLOAT"


class TestExcelParser:
    """Test Excel parser."""

    @pytest.mark.skipif(
        not hasattr(__import__('src.parser.excel_parser', fromlist=['HAS_OPENPYXL']), 'HAS_OPENPYXL'),
        reason="openpyxl not installed"
    )
    def test_excel_parse_simple(self):
        """Test parsing simple Excel file."""
        try:
            from openpyxl import Workbook

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
                wb = Workbook()
                ws = wb.active
                ws.title = "customers"

                # Add header
                ws['A1'] = "id"
                ws['B1'] = "name"
                ws['C1'] = "email"

                # Add data rows
                ws['A2'] = 1
                ws['B2'] = "John"
                ws['C2'] = "john@example.com"

                ws['A3'] = 2
                ws['B3'] = "Jane"
                ws['C3'] = "jane@example.com"

                wb.save(f.name)

                # Test parsing
                parser = ExcelParser()
                with open(f.name, 'rb') as excel_file:
                    content = excel_file.read()

                schema = parser.parse(content)

                assert len(schema.tables) == 1
                table = schema.tables[0]
                assert table.name == "customers"
                assert len(table.columns) == 3
                assert table.estimated_rows == 2

                # Cleanup
                Path(f.name).unlink()

        except ImportError:
            pytest.skip("openpyxl not available")


class TestParserFactory:
    """Test parser factory."""

    def test_factory_detects_csv(self):
        """Test factory detects CSV format."""
        parser = BackupParserFactory.create_parser("data.csv")
        assert isinstance(parser, CsvParser)

    def test_factory_detects_excel(self):
        """Test factory detects Excel formats."""
        parser_xlsx = BackupParserFactory.create_parser("data.xlsx")
        assert isinstance(parser_xlsx, ExcelParser)

        parser_xls = BackupParserFactory.create_parser("data.xls")
        assert isinstance(parser_xls, ExcelParser)

    def test_factory_detects_access(self):
        """Test factory detects Access formats."""
        parser_mdb = BackupParserFactory.create_parser("data.mdb")
        assert isinstance(parser_mdb, AccessParser)

        parser_accdb = BackupParserFactory.create_parser("data.accdb")
        assert isinstance(parser_accdb, AccessParser)

    def test_factory_raises_on_unknown(self):
        """Test factory raises on unsupported format."""
        with pytest.raises(ValueError):
            BackupParserFactory.create_parser("data.xyz")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

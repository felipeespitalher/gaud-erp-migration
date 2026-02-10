"""Excel file parser with multi-sheet support."""
from typing import Optional, List
import re

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from src.parser.backup_parser import DatabaseBackupParser
from src.schema.models import SourceSchema, SourceTable, SourceColumn


class ExcelParser(DatabaseBackupParser):
    """Parse Excel files and extract schema from sheets."""

    def parse(
        self,
        content: bytes,
        selected_tables: Optional[List[str]] = None,
    ) -> SourceSchema:
        """
        Parse Excel content and return schema.

        Args:
            content: Excel file content (as bytes)
            selected_tables: Optional list of sheet names to include.
                           If None, include all sheets.

        Returns:
            SourceSchema: Parsed schema with multiple tables (one per sheet)

        Raises:
            ParseException: If parsing fails or openpyxl not installed
        """
        if not HAS_OPENPYXL:
            raise RuntimeError(
                "openpyxl is required for Excel parsing. "
                "Install with: pip install openpyxl"
            )

        from io import BytesIO

        try:
            wb = load_workbook(BytesIO(content))
        except Exception as e:
            raise RuntimeError(f"Failed to parse Excel file: {str(e)}")

        schema = SourceSchema(database_type="excel", created_at=None)

        # Process each sheet
        for sheet_name in wb.sheetnames:
            # Filter by selected_tables if provided
            if selected_tables and sheet_name not in selected_tables:
                continue

            ws = wb[sheet_name]

            # Extract headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else f"Column_{len(headers)}")

            if not headers:
                continue

            # Count data rows
            row_count = 0
            sample_rows = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_count += 1

                # Collect sample for type inference (first 100 rows)
                if row_idx <= 102:  # First 100 data rows
                    sample_rows.append(list(row))

            # Infer types from sample
            inferred_types = self._infer_types(headers, sample_rows)

            # Create columns
            columns = []
            for col_idx, header in enumerate(headers):
                col = SourceColumn(
                    name=str(header).strip(),
                    type=inferred_types.get(col_idx, "VARCHAR"),
                    nullable=True,
                )
                columns.append(col)

            # Create table
            table = SourceTable(
                name=sheet_name,
                columns=columns,
                estimated_rows=row_count,
            )

            schema.tables.append(table)

        return schema

    def _infer_types(self, headers: List[str], rows: List[List]) -> dict:
        """
        Infer column types from data.

        Args:
            headers: Column headers
            rows: Sample data rows

        Returns:
            dict: {column_index: inferred_type}
        """
        inferred = {}

        for col_idx in range(len(headers)):
            # Default to VARCHAR
            inferred_type = "VARCHAR"

            # Sample values from column
            sample_values = []
            for row in rows:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None:
                        sample_values.append(str(val).strip())

            if sample_values:
                inferred_type = self._infer_type_from_sample(sample_values)

            inferred[col_idx] = inferred_type

        return inferred

    def _infer_type_from_sample(self, values: List[str]) -> str:
        """
        Infer type from sample values.

        Args:
            values: List of string values

        Returns:
            str: Inferred type name
        """
        if not values:
            return "VARCHAR"

        # Count type matches
        is_int_count = 0
        is_float_count = 0
        is_date_count = 0
        is_bool_count = 0

        for val in values:
            val = val.strip()

            # Check if integer
            try:
                int(val)
                is_int_count += 1
                continue
            except ValueError:
                pass

            # Check if float
            try:
                float(val)
                is_float_count += 1
                continue
            except ValueError:
                pass

            # Check if date (simple patterns)
            if re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', val) or \
               re.match(r'^\d{4}-\d{2}-\d{2}', val) or \
               re.match(r'^\d{2}-\d{2}-\d{4}$', val):
                is_date_count += 1
                continue

            # Check if boolean
            if val.lower() in ('true', 'false', 'yes', 'no', '1', '0'):
                is_bool_count += 1
                continue

        # Determine type by majority
        total = len(values)
        threshold = 0.8  # 80% of values must match type

        if is_int_count >= total * threshold:
            return "INTEGER"
        elif is_float_count >= total * threshold:
            return "FLOAT"
        elif is_date_count >= total * threshold:
            return "DATE"
        elif is_bool_count >= total * threshold:
            return "BOOLEAN"
        else:
            return "VARCHAR"
